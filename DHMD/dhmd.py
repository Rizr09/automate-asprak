import re
from pathlib import Path
import PyPDF2
import openpyxl
from openpyxl.styles import Font

def extract_text_from_pdf(pdf_path) -> str:
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text = ""
        for page in reader.pages:
            text += page.extract_text()
    return text

def extract_student_data(text) -> list:
    # Regular expression to match NPM and NAMA Mahasiswa
    pattern = r'(\d{12})\s+(.*?)(?=\s+\.\.\.|\n)'
    matches = re.findall(pattern, text)
    matches = [(npm, nama.title()) for npm, nama in matches]
    return matches

def save_to_excel(data, filename) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Data"

    # Add header
    ws['A1'] = 'NPM'
    ws['B1'] = 'Nama Mahasiswa'
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Add data
    for row, (npm, nama) in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=npm)
        ws.cell(row=row, column=2, value=nama)

    # Adjust column width
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40

    wb.save(filename)

def process_pdf(pdf_path, output_folder) -> None:
    # Extract text from PDF
    pdf_content = extract_text_from_pdf(pdf_path)

    # Extract student data
    student_data = extract_student_data(pdf_content)

    # Generate Excel filename
    excel_filename = output_folder / (pdf_path.stem + '.xlsx')

    # Save data to Excel
    save_to_excel(student_data, excel_filename)

    print(f"Data has been extracted from {pdf_path.name} and saved to {excel_filename}")

def main(input_folder, output_folder):    # Create output folder if it doesn't exist
    output_folder.mkdir(parents=True, exist_ok=True)

    # Process all PDF files in the input folder
    for pdf_file in input_folder.glob('*.pdf'):
        process_pdf(pdf_file, output_folder)

# Set up input and output folders
input_folder = Path('DHMD')
output_folder = Path('DHMD_Excel')

# Execute main function
main(input_folder, output_folder)